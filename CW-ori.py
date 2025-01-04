# any changes made in class SavedStocksManager(ctk.CTkToplevel) no longer works, please update on class StockManager(ctk.CTkToplevel)

import numpy as np
import tkinter as tk
import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox, simpledialog
import pandas as pd
from textblob import TextBlob
import feedparser
import logging
import datetime
import json
from functools import partial
import threading
import requests
import time
from dateutil import parser
from tkcalendar import DateEntry
import os
import subprocess
import re
from html import unescape
from CTkMessagebox import CTkMessagebox
import traceback
import webbrowser
import pyperclip
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import queue
import yfinance as yf
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from stockstats import StockDataFrame
import matplotlib.dates as mdates
import pytz
import mplfinance as mpf
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import io
import gc
import glob
from datetime import datetime, timezone, timedelta
from dateutil import parser
from datetime import datetime, timezone
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout
from PyQt5.QtWebEngineWidgets import QWebEngineView
from PyQt5.QtCore import QUrl
from tradingview_widget import run_tradingview_widget, QApplication
import multiprocessing
from economic_calendar import run_economic_calendar_widget
from tradingchart import run_trading_chart_widget
from bursa_heatmap import run_bursa_heatmap_widget
from us_heatmap import run_us_heatmap_widget
from AI import create_ai_summary_widget
from fear_greed_widget import run_fear_greed_widget  # For Malaysia
from us_fear_greed import run_us_fear_greed_widget
from active_gainers_losers import run_stocks_widget
import streamlit as st
from io import StringIO
from app import main as app_main
from train_stock_model import main
from PyQt5.QtGui import QIcon

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Set appearance mode and color theme
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")




class ProgressBar:
    def __init__(self, master):
        self.master = master
        self.window = ctk.CTkToplevel(master)
        self.window.title("Fetching News")
        self.window.geometry("400x150")
        self.window.transient(master)
        self.window.grab_set()
        self.is_alive = True

        self.progress = ctk.CTkProgressBar(self.window, mode='indeterminate', width=300)
        self.progress.pack(pady=(20, 10))

        self.status_label = ctk.CTkLabel(self.window, text="Initializing...")
        self.status_label.pack(pady=5)

        self.feed_label = ctk.CTkLabel(self.window, text="")
        self.feed_label.pack(pady=5)

        self.item_label = ctk.CTkLabel(self.window, text="")
        self.item_label.pack(pady=5)

        self.window.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        self.is_alive = False
        if self.master.winfo_exists():
            self.window.destroy()

    def start(self):
        self.progress.start()

    def stop(self):
        self.is_alive = False
        if self.master.winfo_exists():
            self.window.after(0, self.window.destroy)

    def update_status(self, status):
        if self.is_alive and self.master.winfo_exists():
            self.window.after(0, lambda: self.status_label.configure(text=status))

    def update_feed(self, feed_name):
        if self.is_alive and self.master.winfo_exists():
            self.window.after(0, lambda: self.feed_label.configure(text=f"Current feed: {feed_name}"))

    def update_item_count(self, count, total):
        if self.is_alive and self.master.winfo_exists():
            self.window.after(0, lambda: self.item_label.configure(text=f"Processed items: {count}/{total}"))



class StockPriceViewer(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Stock Price Viewer")
        self.geometry("1000x500")
        self.minsize(500, 600)
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self.api_choice = tk.StringVar(value="yfinance")
        self.alpha_vantage_key = self.load_alpha_vantage_key()
        self.period = tk.StringVar(value="1mo")
        self.symbol = ""
        self.saved_stocks = self.load_saved_stocks()
        
        self.create_widgets()
        
        self.transient(parent)
        self.grab_set()

        # Set up logging
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler('stock_viewer.log')
        fh.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        self.logger.addHandler(fh)

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_widgets(self):
        # Input frame
        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        input_frame.grid_columnconfigure(7, weight=1)

        ctk.CTkLabel(input_frame, text="Stock Symbol:").grid(row=0, column=0, padx=5, pady=5)
        self.stock_entry = ctk.CTkEntry(input_frame, width=100)
        self.stock_entry.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(input_frame, text="Exchange:").grid(row=0, column=2, padx=5, pady=5)
        self.exchange_var = tk.StringVar(value="US")
        exchange_menu = ctk.CTkOptionMenu(input_frame, variable=self.exchange_var, values=["US", "MY"])
        exchange_menu.grid(row=0, column=3, padx=5, pady=5)

        ctk.CTkButton(input_frame, text="Get Stock Info", command=self.get_stock_info_threaded).grid(row=0, column=4, padx=5, pady=5)

        ctk.CTkLabel(input_frame, text="Period:").grid(row=0, column=5, padx=5, pady=5)
        period_menu = ctk.CTkOptionMenu(input_frame, variable=self.period, 
                                        values=["1d", "5d", "1mo", "6mo", "1y"],
                                        command=self.update_chart)
        period_menu.grid(row=0, column=6, padx=5, pady=5)

        ctk.CTkButton(input_frame, text="Manage Saved Stocks", command=self.manage_saved_stocks).grid(row=0, column=7, padx=5, pady=5, sticky="e")

        # Info frame
        self.info_frame = ctk.CTkFrame(self)
        self.info_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        self.info_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        # Chart frame
        self.chart_frame = ctk.CTkFrame(self)
        self.chart_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        self.chart_frame.grid_columnconfigure(0, weight=1)
        self.chart_frame.grid_rowconfigure(0, weight=1)

        
        
    def load_saved_stocks(self):
        try:
            with open('saved_stocks.json', 'r') as f:
                content = f.read()
            
            # Attempt to fix common JSON errors
            content = re.sub(r'}\s*{', '},{', content)  # Add missing commas
            content = re.sub(r',\s*]', ']', content)    # Remove trailing commas
            content = content.replace("'", '"')         # Replace single quotes with double quotes
            
            # Wrap the content in square brackets if not already
            if not content.strip().startswith('['):
                content = '[' + content + ']'
            
            # Try to parse the corrected JSON
            stocks = json.loads(content)
            
            # Ensure the result is a list
            if not isinstance(stocks, list):
                stocks = [stocks]
            
            return stocks
        except FileNotFoundError:
            print("saved_stocks.json file not found. Creating a new one.")
            return []
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON: {e}")
            print("Please check your saved_stocks.json file for formatting errors.")
            return []
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            return []


    def format_my_stock_symbol(symbol):
        if symbol.endswith('.KL'):
            base = symbol[:-3]
            return f"{int(base):04d}.KL"
        return symbol


    def save_stocks(self):
        try:
            with open('saved_stocks.json', 'w') as f:
                json.dump(self.saved_stocks, f, indent=2)
            self.logger.info("Saved stocks updated in JSON file")
        except Exception as e:
            self.logger.error(f"Error saving stocks to JSON: {str(e)}")
            CTkMessagebox(title="Error", message=f"Failed to save stocks: {str(e)}", icon="cancel")

    def manage_saved_stocks(self):
        stock_manager = StockManager(self)
        self.wait_window(stock_manager)  # Wait for the window to be closed
        if stock_manager.selected_symbol:
            self.stock_entry.delete(0, tk.END)
            self.stock_entry.insert(0, stock_manager.selected_symbol)


    def set_alpha_vantage_key(self):
        current_key = self.alpha_vantage_key if self.alpha_vantage_key else "No key set"
        key = ctk.CTkInputDialog(title="Alpha Vantage API Key", text=f"Current key: {current_key}\nEnter new Alpha Vantage API Key:").get_input()
        if key:
            self.alpha_vantage_key = key
            self.save_alpha_vantage_key(key)
            CTkMessagebox(title="Success", message="Alpha Vantage API Key set and saved successfully.", icon="info")


    def save_alpha_vantage_key(self, key):
        with open('alpha_vantage_key.json', 'w') as f:
            json.dump({"key": key}, f)

    def load_alpha_vantage_key(self):
        try:
            with open('alpha_vantage_key.json', 'r') as f:
                data = json.load(f)
                return data.get("key", "")
        except FileNotFoundError:
            return ""
            
    def get_stock_info_threaded(self):
        self.logger.info("Starting get_stock_info_threaded")
        threading.Thread(target=self.get_stock_info, daemon=True).start()

    def get_stock_info(self):
        self.logger.info("Starting get_stock_info")
        self.symbol = self.stock_entry.get().strip().upper()
        exchange = self.exchange_var.get()

        if not self.symbol:
            self.logger.warning("No symbol entered")
            self.after(0, lambda: CTkMessagebox(title="Error", message="Please enter a stock symbol.", icon="warning"))
            return

        # For Malaysian stocks, ensure the symbol has 4 digits and .KL suffix
        if exchange == "MY":
            if self.symbol.isdigit():
                self.symbol = f"{int(self.symbol):04d}.KL"
            elif not self.symbol.endswith('.KL'):
                self.symbol += '.KL'
            self.stock_entry.delete(0, tk.END)
            self.stock_entry.insert(0, self.symbol)

        try:
            self.get_yfinance_info(self.symbol)
        except Exception as e:
            self.logger.error(f"Error in get_stock_info: {str(e)}")
            self.logger.error(traceback.format_exc())
            self.after(0, lambda: CTkMessagebox(title="Error", message=f"An error occurred: {str(e)}", icon="cancel"))
            
    def update_chart(self, *args):
        if self.symbol:
            self.get_yfinance_info(self.symbol, plot_only=True)

    def get_yfinance_info(self, symbol, plot_only=False):
        self.logger.info(f"Starting get_yfinance_info for {symbol}")
        try:
            stock = yf.Ticker(symbol)
            info = stock.info

            if not plot_only:
                self.after(0, lambda: self.display_stock_info(info))

            period = self.period.get()
            end = datetime.now(pytz.timezone('US/Eastern'))

            if period == '1d':
                start = end - timedelta(days=1)
                start = start.replace(hour=9, minute=30, second=0, microsecond=0)
                end = end.replace(hour=16, minute=0, second=0, microsecond=0)
                
                if end.weekday() >= 5:  # If it's weekend
                    end = end - timedelta(days=end.weekday() - 4)  # Set to last Friday
                    start = end - timedelta(days=1)
                    start = start.replace(hour=9, minute=30, second=0, microsecond=0)
                    end = end.replace(hour=16, minute=0, second=0, microsecond=0)
                
                history = stock.history(start=start, end=end, interval="1m")
            elif period == '5d':
                history = stock.history(period="5d", interval="1h")
            else:  # '1mo', '6mo', '1y'
                history = stock.history(period=period, interval="1d")

            if history.empty:
                self.logger.warning(f"No data available for {symbol} in period {period}")
                self.after(0, lambda: CTkMessagebox(title="No Data", message=f"No data available for the selected period ({period}).", icon="warning"))
                return

            self.after(0, lambda: self.plot_stock_price_plotly(history, symbol, info))

        except Exception as e:
            self.logger.error(f"Error in get_yfinance_info: {str(e)}")
            self.logger.error(traceback.format_exc())
            error_message = f"Failed to fetch stock information: {str(e)}"
            self.after(0, lambda: CTkMessagebox(title="Error", message=error_message, icon="cancel"))


    def calculate_pvsra(self, data):
        # Calculate average volume for the last 10 candles
        data['avg_volume'] = data['Volume'].rolling(window=10).mean()
        
        # Calculate volume surge (200% or more of average)
        data['volume_surge'] = data['Volume'] >= 2 * data['avg_volume']
        
        # Calculate candle spread
        data['candle_spread'] = abs(data['Close'] - data['Open'])
        
        # Calculate spread * volume product
        data['spread_volume_product'] = data['candle_spread'] * data['Volume']
        
        # Calculate the highest spread * volume product for the last 10 candles
        data['max_spread_volume'] = data['spread_volume_product'].rolling(window=10).max()
        
        # Identify PVSRA candles
        data['pvsra_bull'] = (data['volume_surge'] & 
                              (data['spread_volume_product'] >= data['max_spread_volume']) & 
                              (data['Close'] > data['Open']))
        
        data['pvsra_bear'] = (data['volume_surge'] & 
                              (data['spread_volume_product'] >= data['max_spread_volume']) & 
                              (data['Close'] <= data['Open']))
        
        return data


    def get_alpha_vantage_info(self, symbol):
        if not self.alpha_vantage_key:
            self.after(0, lambda: CTkMessagebox(title="Error", message="Please set your Alpha Vantage API Key first.", icon="warning"))
            return

        try:
            # Get company overview
            overview_url = f"https://www.alphavantage.co/query?function=OVERVIEW&symbol={symbol}&apikey={self.alpha_vantage_key}"
            overview_response = requests.get(overview_url)
            overview_response.raise_for_status()
            overview_data = overview_response.json()
            logging.debug(f"Overview data: {overview_data}")

            # Check if we've hit the API call limit
            if "Information" in overview_data and "standard API rate limit" in overview_data["Information"]:
                logging.warning("Alpha Vantage API limit reached. Falling back to yfinance.")
                self.after(0, lambda: CTkMessagebox(title="API Limit Reached", 
                                                    message="Alpha Vantage API limit reached. Falling back to yfinance data.",
                                                    icon="warning"))
                self.get_yfinance_info(symbol)
                return

            # If we haven't hit the limit, proceed with Alpha Vantage data
            quote_url = f"https://www.alphavantage.co/query?function=GLOBAL_QUOTE&symbol={symbol}&apikey={self.alpha_vantage_key}"
            quote_response = requests.get(quote_url)
            quote_response.raise_for_status()
            quote_data = quote_response.json()
            logging.debug(f"Quote data: {quote_data}")

            # Combine data and display stock information
            combined_info = {**overview_data, **quote_data.get("Global Quote", {})}
            self.after(0, lambda: self.display_stock_info(combined_info))

            # Get time series data for plotting
            time_series_url = f"https://www.alphavantage.co/query?function=TIME_SERIES_DAILY&symbol={symbol}&apikey={self.alpha_vantage_key}"
            time_series_response = requests.get(time_series_url)
            time_series_response.raise_for_status()
            time_series_data = time_series_response.json()
            logging.debug(f"Time series data keys: {time_series_data.keys()}")

            if "Time Series (Daily)" in time_series_data:
                df = pd.DataFrame(time_series_data["Time Series (Daily)"]).T
                df.index = pd.to_datetime(df.index)
                df = df.astype(float)
                df.columns = ['Open', 'High', 'Low', 'Close', 'Volume']  # Rename columns
                df = df.sort_index()  # Sort by date
                logging.debug(f"Dataframe head: {df.head()}")
                logging.debug(f"Dataframe columns: {df.columns}")
                self.after(0, lambda: self.plot_stock_price(df))
            else:
                error_message = "Unable to fetch historical data for plotting from Alpha Vantage."
                logging.error(f"{error_message} Available keys: {time_series_data.keys()}")
                self.after(0, lambda: CTkMessagebox(title="Error", message=error_message, icon="warning"))
                # Fallback to yfinance for plotting
                self.get_yfinance_info(symbol, plot_only=True)

        except requests.exceptions.RequestException as e:
            error_message = f"Network error when fetching stock information: {str(e)}"
            logging.error(error_message)
            self.after(0, lambda: CTkMessagebox(title="Network Error", message=error_message, icon="cancel"))
            # Fallback to yfinance
            self.get_yfinance_info(symbol)
        except Exception as e:
            error_message = f"Failed to fetch stock information: {str(e)}"
            logging.error(error_message)
            logging.error(traceback.format_exc())
            self.after(0, lambda: CTkMessagebox(title="Error", message=error_message, icon="cancel"))
            # Fallback to yfinance
            self.get_yfinance_info(symbol)        
        
    def display_stock_info(self, info):
        for widget in self.info_frame.winfo_children():
            widget.destroy()

        self.info_frame.grid_columnconfigure(tuple(range(4)), weight=1)

        labels = [
            ("Symbol", "symbol"), ("Company Name", "longName"), ("Current Price", "currentPrice"),
            ("Previous Close", "previousClose"), ("Open", "open"), ("Day's Range", "dayRange"),
            ("52 Week Range", "52WeekRange"), ("Volume", "volume"), ("Avg. Volume", "averageVolume"),
            ("Market Cap", "marketCap"), ("Beta", "beta"), ("P/E Ratio", "trailingPE"),
            ("EPS", "trailingEps"), ("Forward Dividend & Yield", "dividendYield"),
            ("Ex-Dividend Date", "exDividendDate"), ("1y Target Est", "targetMeanPrice"),
        ]

        for i, (label, key) in enumerate(labels):
            value = info.get(key, "N/A")
            if isinstance(value, (int, float)):
                value = f"{value:.2f}"
            ctk.CTkLabel(self.info_frame, text=f"{label}: {value}", anchor="w").grid(row=i//4, column=i%4, sticky="w", padx=5, pady=2)


    def format_dividend_info(self, info):
        dividend_rate = info.get('dividendRate', 'N/A')
        dividend_yield = info.get('dividendYield', 'N/A')
        
        if isinstance(dividend_rate, (int, float)):
            dividend_rate = f"${dividend_rate:.2f}"
        
        if isinstance(dividend_yield, (int, float)):
            dividend_yield = f"{dividend_yield:.2%}"
        
        return f"Div & Yield: {dividend_rate} ({dividend_yield})"


    def plot_stock_price_plotly(self, data, symbol, info):
        try:
            self.logger.info(f"Starting plot_stock_price_plotly for {symbol}")
            
            def format_dividend_info(info):
                dividend_rate = info.get('dividendRate', 'N/A')
                dividend_yield = info.get('dividendYield', 'N/A')
                
                if isinstance(dividend_rate, (int, float)):
                    dividend_rate = f"${dividend_rate:.2f}"
                
                if isinstance(dividend_yield, (int, float)):
                    dividend_yield = f"{dividend_yield:.2%}"
                
                return f"Div & Yield: {dividend_rate} ({dividend_yield})"

            def generate_chart():
                try:
                    self.logger.info("Calculating indicators")
                    data_with_indicators = self.calculate_indicators(data)
                    
                    # Calculate PVSRA indicators
                    data_with_indicators = self.calculate_pvsra(data_with_indicators)

                    # Determine the correct interval for the title
                    interval_text = {
                        '1d': '1-Minute Intervals',
                        '5d': '1-Hour Intervals',
                        '1mo': 'Daily Intervals',
                        '6mo': 'Daily Intervals',
                        '1y': 'Daily Intervals'
                    }.get(self.period.get(), 'Custom Intervals')

                    self.logger.info("Creating subplots")
                    fig = make_subplots(
                        rows=7, cols=1,
                        shared_xaxes=True,
                        vertical_spacing=0.03,
                        row_heights=[0.4, 0.1, 0.1, 0.1, 0.1, 0.1, 0.1],
                        subplot_titles=(f"Price ({interval_text})", "Volume", "MACD", "Stochastic Oscillator", "RSI", "ADX", "Bollinger Bands %B")
                    )

                    self.logger.info("Adding candlestick chart")
                    fig.add_trace(go.Candlestick(
                        x=data_with_indicators.index,
                        open=data_with_indicators['Open'],
                        high=data_with_indicators['High'],
                        low=data_with_indicators['Low'],
                        close=data_with_indicators['Close'],
                        name='Price',
                        increasing_line_color='green',  # Color for bullish candles
                        decreasing_line_color='red'     # Color for bearish candles
                    ), row=1, col=1)


                    # Add EMAs with distinct colors
                    ema_colors = ['blue', 'green', 'red']
                    for i, period in enumerate([10, 20, 50]):
                        fig.add_trace(go.Scatter(
                            x=data_with_indicators.index,
                            y=data_with_indicators[f'EMA_{period}'],
                            name=f'EMA {period}',
                            line=dict(width=1, color=ema_colors[i])
                        ), row=1, col=1)

                    # Add Bollinger Bands
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Upper_BB'],
                        name='Upper BB',
                        line=dict(color='rgba(250, 0, 0, 0.5)')
                    ), row=1, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Lower_BB'],
                        name='Lower BB',
                        line=dict(color='rgba(0, 250, 0, 0.5)')
                    ), row=1, col=1)

                    # Add Ichimoku Cloud
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Tenkan_Sen'],
                        name='Tenkan Sen',
                        line=dict(color='purple', width=1)
                    ), row=1, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Kijun_Sen'],
                        name='Kijun Sen',
                        line=dict(color='brown', width=1)
                    ), row=1, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Senkou_Span_A'],
                        name='Senkou Span A',
                        line=dict(color='rgba(0,255,0,0.3)', width=1)
                    ), row=1, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Senkou_Span_B'],
                        name='Senkou Span B',
                        line=dict(color='rgba(255,0,0,0.3)', width=1),
                        fill='tonexty'
                    ), row=1, col=1)

                    # Add Buy/Sell Signals
                    bb_buy_signals = data_with_indicators[data_with_indicators['BB_Buy_Signal']]
                    bb_sell_signals = data_with_indicators[data_with_indicators['BB_Sell_Signal']]
                    ichimoku_buy = data_with_indicators[data_with_indicators['Ichimoku_Buy_Signal']]
                    ichimoku_sell = data_with_indicators[data_with_indicators['Ichimoku_Sell_Signal']]
                    
                    fig.add_trace(go.Scatter(
                        x=bb_buy_signals.index,
                        y=bb_buy_signals['Low'],
                        mode='markers',
                        name='Bollinger Buy Signal',
                        marker=dict(symbol='triangle-up', size=10, color='green')
                    ), row=1, col=1)
                    
                    fig.add_trace(go.Scatter(
                        x=bb_sell_signals.index,
                        y=bb_sell_signals['High'],
                        mode='markers',
                        name='Bollinger Sell Signal',
                        marker=dict(symbol='triangle-down', size=10, color='red')
                    ), row=1, col=1)
                    
                    fig.add_trace(go.Scatter(
                        x=ichimoku_buy.index,
                        y=ichimoku_buy['Low'],
                        mode='markers',
                        name='Ichimoku Buy Signal',
                        marker=dict(symbol='triangle-up', size=10, color='cyan')
                    ), row=1, col=1)
                    
                    fig.add_trace(go.Scatter(
                        x=ichimoku_sell.index,
                        y=ichimoku_sell['High'],
                        mode='markers',
                        name='Ichimoku Sell Signal',
                        marker=dict(symbol='triangle-down', size=10, color='magenta')
                    ), row=1, col=1)

                    # Add PVSRA indicators
                    pvsra_bull = data_with_indicators[data_with_indicators['pvsra_bull']]
                    pvsra_bear = data_with_indicators[data_with_indicators['pvsra_bear']]

                    fig.add_trace(go.Scatter(
                        x=pvsra_bull.index,
                        y=pvsra_bull['Low'],
                        mode='markers',
                        name='PVSRA Bullish',
                        marker=dict(symbol='triangle-up', size=10, color='darkgreen'),
                        hoverinfo='text',
                        text='PVSRA Bullish'
                    ), row=1, col=1)

                    fig.add_trace(go.Scatter(
                        x=pvsra_bear.index,
                        y=pvsra_bear['High'],
                        mode='markers',
                        name='PVSRA Bearish',
                        marker=dict(symbol='triangle-down', size=10, color='darkred'),
                        hoverinfo='text',
                        text='PVSRA Bearish'
                    ), row=1, col=1)

                    # Add Hammer indicator
                    hammer_signals = data_with_indicators[data_with_indicators['Hammer']]
                    fig.add_trace(go.Scatter(
                        x=hammer_signals.index,
                        y=hammer_signals['Low'],
                        mode='markers',
                        name='Hammer',
                        marker=dict(symbol='triangle-up', size=10, color='orange'),
                        hoverinfo='text',
                        text='Hammer'
                    ), row=1, col=1)

                    # Add Engulfing Patterns
                    bullish_engulfing = data_with_indicators[data_with_indicators['Bullish_Engulfing']]
                    bearish_engulfing = data_with_indicators[data_with_indicators['Bearish_Engulfing']]

                    fig.add_trace(go.Scatter(
                        x=bullish_engulfing.index,
                        y=bullish_engulfing['Low'],
                        mode='markers',
                        name='Bullish Engulfing',
                        marker=dict(symbol='triangle-up', size=10, color='lime'),
                        hoverinfo='text',
                        text='Bullish Engulfing'
                    ), row=1, col=1)

                    fig.add_trace(go.Scatter(
                        x=bearish_engulfing.index,
                        y=bearish_engulfing['High'],
                        mode='markers',
                        name='Bearish Engulfing',
                        marker=dict(symbol='triangle-down', size=10, color='pink'),
                        hoverinfo='text',
                        text='Bearish Engulfing'
                    ), row=1, col=1)

                    # Add Marubozu Candles
                    marubozu_candles = data_with_indicators[data_with_indicators['Marubozu']]
                    fig.add_trace(go.Scatter(
                        x=marubozu_candles.index,
                        y=marubozu_candles['High'],
                        mode='markers',
                        name='Marubozu Candle',
                        marker=dict(symbol='star', size=10, color='gold'),
                        hoverinfo='text',
                        text='Marubozu Candle'
                    ), row=1, col=1)

                    # Volume chart
                    fig.add_trace(go.Bar(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Volume'],
                        name='Volume',
                        marker_color='lightblue'
                    ), row=2, col=1)

                    # MACD
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['MACD'],
                        name='MACD Line',
                        line=dict(color='blue')
                    ), row=3, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Signal'],
                        name='Signal Line',
                        line=dict(color='orange')
                    ), row=3, col=1)
                    fig.add_trace(go.Bar(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Histogram'],
                        name='MACD Histogram',
                        marker_color='gray'
                    ), row=3, col=1)

                    # Stochastic Oscillator
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['%K'],
                        name='Fast %K (14-period)',
                        line=dict(color='blue')
                    ), row=4, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['%D'],
                        name='Slow %K (3-period SMA of Fast %K)',
                        line=dict(color='red')
                    ), row=4, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Slow %D'],
                        name='Slow %D (3-period SMA of Slow %K)',
                        line=dict(color='green')
                    ), row=4, col=1)

                    # RSI
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['RSI'],
                        name='RSI',
                        line=dict(color='purple')
                    ), row=5, col=1)

                    # ADX
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['ADX'],
                        name='ADX',
                        line=dict(color='black')
                    ), row=6, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Plus_DI'],
                        name='+DI',
                        line=dict(color='green')
                    ), row=6, col=1)
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['Minus_DI'],
                        name='-DI',
                        line=dict(color='red')
                    ), row=6, col=1)

                    # Bollinger Bands %B
                    data_with_indicators['BB_%B'] = (data_with_indicators['Close'] - data_with_indicators['Lower_BB']) / (data_with_indicators['Upper_BB'] - data_with_indicators['Lower_BB'])
                    fig.add_trace(go.Scatter(
                        x=data_with_indicators.index,
                        y=data_with_indicators['BB_%B'],
                        name='Bollinger Bands %B',
                        line=dict(color='darkblue')
                    ), row=7, col=1)

                    self.logger.info("Updating layout")
                    fig.update_layout(
                        title=f'{symbol} Stock Price History ({self.period.get()}, {interval_text})',
                        height=2000,
                        width=1200,
                        showlegend=True,
                        legend=dict(
                            yanchor="top",
                            y=0.99,
                            xanchor="left",
                            x=1.05,
                            traceorder="grouped",
                            itemsizing="constant"
                        ),
                        plot_bgcolor='white',
                        paper_bgcolor='white'
                    )

                    # Adjust x-axis date formatting based on the selected period
                    if self.period.get() == '1d':
                        fig.update_xaxes(
                            rangeslider_visible=False,
                            rangeselector=dict(visible=False),
                            tickformat='%H:%M',
                            tickangle=45,
                            tickmode='auto',
                            nticks=20
                        )
                    elif self.period.get() == '5d':
                        fig.update_xaxes(
                            rangeslider_visible=False,
                            rangeselector=dict(visible=False),
                            tickformat='%m/%d %H:%M',
                            tickangle=45,
                            tickmode='auto',
                            nticks=20
                        )
                    else:
                        fig.update_xaxes(
                            rangeslider_visible=False,
                            rangeselector=dict(visible=False),
                            tickformat='%Y-%m-%d',
                            tickangle=45,
                            tickmode='auto',
                            nticks=20
                        )

                    # Update y-axis labels and colors
                    fig.update_yaxes(title_text="Price", row=1, col=1, gridcolor="lightgray", tickfont=dict(color="black"))
                    fig.update_yaxes(title_text="Volume", row=2, col=1, gridcolor="lightgray", tickfont=dict(color="black"))
                    fig.update_yaxes(title_text="MACD", row=3, col=1, gridcolor="lightgray", tickfont=dict(color="black"))
                    fig.update_yaxes(title_text="Stochastic", row=4, col=1, gridcolor="lightgray", tickfont=dict(color="black"))
                    fig.update_yaxes(title_text="RSI", row=5, col=1, gridcolor="lightgray", tickfont=dict(color="black"))
                    fig.update_yaxes(title_text="ADX", row=6, col=1, gridcolor="lightgray", tickfont=dict(color="black"))
                    fig.update_yaxes(title_text="BB %B", row=7, col=1, gridcolor="lightgray", tickfont=dict(color="black"))

                    # Group legend items
                    fig.update_layout(legend_groupclick="toggleitem")

                    # Add horizontal lines for overbought/oversold levels
                    fig.add_shape(type="line", x0=0, x1=1, y0=80, y1=80,
                                  line=dict(color="red", width=2, dash="dash"),
                                  xref="paper", yref="y4")
                    fig.add_shape(type="line", x0=0, x1=1, y0=20, y1=20,
                                  line=dict(color="green", width=2, dash="dash"),
                                  xref="paper", yref="y4")
                    fig.add_shape(type="line", x0=0, x1=1, y0=70, y1=70,
                                  line=dict(color="red", width=2, dash="dash"),
                                  xref="paper", yref="y5")
                    fig.add_shape(type="line", x0=0, x1=1, y0=30, y1=30,
                                  line=dict(color="green", width=2, dash="dash"),
                                  xref="paper", yref="y5")
                    fig.add_shape(type="line", x0=0, x1=1, y0=25, y1=25,
                                  line=dict(color="gray", width=2, dash="dash"),
                                  xref="paper", yref="y6")
                    fig.add_shape(type="line", x0=0, x1=1, y0=1, y1=1,
                                  line=dict(color="red", width=2, dash="dash"),
                                  xref="paper", yref="y7")
                    fig.add_shape(type="line", x0=0, x1=1, y0=0, y1=0,
                                  line=dict(color="green", width=2, dash="dash"),
                                  xref="paper", yref="y7")

                    # Make cursor consistent across all subplots
                    fig.update_layout(hovermode="x unified")

                    # Create a separate table for stock information with two columns
                    left_column = [
                        f"Company: {info.get('longName', 'N/A')}",
                        f"Symbol: {symbol}",
                        f"Current: ${info.get('currentPrice', 'N/A')}",
                        f"Prev Close: ${info.get('previousClose', 'N/A')}",
                        f"Open: ${info.get('open', 'N/A')}",
                        f"Day Range: ${info.get('dayLow', 'N/A')} - ${info.get('dayHigh', 'N/A')}",
                        f"52W Range: ${info.get('fiftyTwoWeekLow', 'N/A')} - ${info.get('fiftyTwoWeekHigh', 'N/A')}"
                    ]
                    
                    right_column = [
                        f"Volume: {info.get('volume', 'N/A')}",
                        f"Avg Vol: {info.get('averageVolume', 'N/A')}",
                        f"Market Cap: ${info.get('marketCap', 'N/A')}",
                        f"Beta: {info.get('beta', 'N/A')}",
                        f"P/E: {info.get('trailingPE', 'N/A')}",
                        f"EPS: ${info.get('trailingEps', 'N/A')}",
                        self.format_dividend_info(info),
                        f"1y Target: ${info.get('targetMeanPrice', 'N/A')}"
                    ]

                    info_table = go.Figure(data=[go.Table(
                        header=dict(values=["Stock Information", ""],
                                    fill_color='paleturquoise',
                                    align='left'),
                        cells=dict(values=[left_column, right_column],
                                   fill_color='lavender',
                                   align='left')
                    )])

                    info_table.update_layout(
                        height=250,
                        width=1200,
                        margin=dict(l=0, r=0, t=30, b=0)
                    )

                    self.logger.info("Saving charts as HTML")
                    filename = f"chart_{symbol}_{int(time.time())}.html"
                    with open(filename, 'w') as f:
                        f.write(info_table.to_html(full_html=False, include_plotlyjs='cdn'))
                        f.write(fig.to_html(full_html=False, include_plotlyjs='cdn'))
                    
                    self.logger.info(f"Charts generated and saved as {filename}")
                    return filename
                except Exception as e:
                    self.logger.error(f"Error in generate_chart: {str(e)}")
                    self.logger.error(traceback.format_exc())
                    return None

            # Show a progress dialog
            progress_dialog = CTkMessagebox(title="Generating Chart", 
                                            message="Please wait while the chart is being generated...",
                                            icon="info")
            self.update()  # Force update of the GUI

            try:
                # Generate the chart
                self.logger.info("Starting chart generation")
                filename = generate_chart()

                # Close the progress dialog
                progress_dialog.destroy()

                if filename:
                    self.logger.info(f"Chart generated successfully: {filename}")
                    self.display_chart(filename)
                else:
                    self.logger.warning("Chart generation failed, displaying error message")
                    CTkMessagebox(title="Error", message="Failed to generate chart. Please check the logs for details.", icon="cancel")
            except Exception as e:
                self.logger.error(f"Unexpected error in plot_stock_price_plotly: {str(e)}")
                self.logger.error(traceback.format_exc())
                progress_dialog.destroy()
                CTkMessagebox(title="Error", message=f"An unexpected error occurred: {str(e)}", icon="cancel")
        except Exception as e:
            self.logger.error(f"Error in plot_stock_price_plotly: {str(e)}")
            self.logger.error(traceback.format_exc())
            raise


            
    def create_stock_info_table(self, symbol, info):
        try:
            self.logger.info("Creating stock info table")
            # Define the information we want to display
            left_column = [
                f"Company: {info.get('longName', 'N/A')}",
                f"Symbol: {symbol}",
                f"Current: ${info.get('currentPrice', 'N/A'):.2f}",
                f"Prev Close: ${info.get('previousClose', 'N/A'):.2f}",
                f"Open: ${info.get('open', 'N/A'):.2f}",
                f"Day Range: ${info.get('dayLow', 'N/A'):.2f} - ${info.get('dayHigh', 'N/A'):.2f}",
                f"52W Range: ${info.get('fiftyTwoWeekLow', 'N/A'):.2f} - ${info.get('fiftyTwoWeekHigh', 'N/A'):.2f}"
            ]

            right_column = [
                f"Volume: {info.get('volume', 'N/A'):,}",
                f"Avg Vol: {info.get('averageVolume', 'N/A'):,}",
                f"Market Cap: ${info.get('marketCap', 'N/A'):,.0f}",
                f"Beta: {info.get('beta', 'N/A'):.3f}",
                f"P/E: {info.get('trailingPE', 'N/A'):.2f}",
                f"EPS: ${info.get('trailingEps', 'N/A'):.2f}",
                f"Div & Yield: ${info.get('dividendRate', 'N/A'):.2f} ({info.get('dividendYield', 'N/A'):.2%})",
                f"1y Target: ${info.get('targetMeanPrice', 'N/A'):.2f}"
            ]

            table = go.Table(
                header=dict(
                    values=["<b>Stock Information</b>", ""],
                    fill_color='paleturquoise',
                    align='left',
                    font=dict(size=14)
                ),
                cells=dict(
                    values=[left_column, right_column],
                    fill_color=['lavender', 'lavender'],
                    align='left',
                    font=dict(size=12),
                    height=30
                )
            )

            return table
        except Exception as e:
            self.logger.error(f"Error creating stock info table: {str(e)}")
            self.logger.error(traceback.format_exc())
            raise


    def create_stock_info_text(self, symbol, info):
        # Define the information we want to display
        keys = [
            'symbol', 'longName', 'sector', 'industry',
            'currentPrice', 'previousClose', 'open', 'dayLow', 'dayHigh',
            'volume', 'averageVolume', 'marketCap',
            'beta', 'trailingPE', 'forwardPE', 'trailingEps',
            'dividendYield', 'exDividendDate',
            'fiftyTwoWeekLow', 'fiftyTwoWeekHigh'
        ]

        # Create three columns of data
        columns = [[], [], []]
        for i, key in enumerate(keys):
            value = info.get(key, 'N/A')
            if isinstance(value, (int, float)):
                value = f"{value:,.2f}"
            item = f"<b>{key}:</b> {value}"
            columns[i % 3].append(item)

        # Join the columns with HTML line breaks and spaces
        info_text = (
            "   ".join([" | ".join(col) for col in columns])
        ).replace(" | ", "&nbsp;&nbsp;&nbsp;|&nbsp;&nbsp;&nbsp;").replace("\n", "<br>")

        return info_text

      
        
    def open_chart(self, filename):
        with self.chart_lock:
            if len(self.open_charts) >= self.max_open_charts:
                oldest_chart = self.open_charts.pop(0)
                try:
                    os.remove(oldest_chart)
                    self.logger.info(f"Removed old chart: {oldest_chart}")
                except Exception as e:
                    self.logger.error(f"Error removing old chart {oldest_chart}: {str(e)}")

            self.open_charts.append(filename)

        self.logger.info(f"Opening chart: {filename}")
        try:
            webbrowser.open('file://' + os.path.realpath(filename), new=2)
        except Exception as e:
            self.logger.error(f"Error opening chart in browser: {str(e)}")
            CTkMessagebox(title="Error", message=f"Failed to open chart in browser: {str(e)}", icon="cancel")



    def cleanup_old_charts(self):
        current_time = time.time()
        for file in glob.glob("combined_stock_chart_*.html"):
            if file not in self.open_charts:
                file_creation_time = os.path.getctime(file)
                if current_time - file_creation_time > 3600:  # Remove files older than 1 hour
                    try:
                        os.remove(file)
                        self.logger.info(f"Removed old chart file: {file}")
                    except Exception as e:
                        self.logger.error(f"Error removing old chart file {file}: {str(e)}")


        
    def display_chart(self, filename):
        try:
            self.logger.info(f"Attempting to open chart: {filename}")
            webbrowser.open('file://' + os.path.realpath(filename), new=2)
            self.logger.info("Chart opened successfully")
        except Exception as e:
            self.logger.error(f"Error opening chart: {str(e)}")
            CTkMessagebox(title="Error", message=f"Failed to open chart: {str(e)}", icon="cancel")
            
            
    def on_closing(self):
        # Clean up chart files when closing the viewer
        for file in os.listdir():
            if file.startswith("chart_") and file.endswith(".html"):
                try:
                    os.remove(file)
                    self.logger.info(f"Removed chart file on closing: {file}")
                except Exception as e:
                    self.logger.error(f"Error removing chart file {file} on closing: {str(e)}")
        
        self.destroy()

    def open_in_browser(self, filename):
        try:
            webbrowser.open('file://' + os.path.realpath(filename), new=2)
            self.logger.info(f"Chart opened in default web browser: {filename}")
        except Exception as e:
            self.logger.error(f"Error opening chart in web browser: {str(e)}")
            CTkMessagebox(title="Error", message=f"Failed to open chart in web browser: {str(e)}", icon="cancel")
            
            
        
    def calculate_indicators(self, data):
        try:
            self.logger.info("Calculating indicators")
            # Calculate EMAs
            for period in [10, 20, 50]:
                data[f'EMA_{period}'] = data['Close'].ewm(span=period, adjust=False).mean()

            # Calculate MACD
            data['EMA_12'] = data['Close'].ewm(span=12, adjust=False).mean()
            data['EMA_26'] = data['Close'].ewm(span=26, adjust=False).mean()
            data['MACD'] = data['EMA_12'] - data['EMA_26']
            data['Signal'] = data['MACD'].ewm(span=9, adjust=False).mean()
            data['Histogram'] = data['MACD'] - data['Signal']

            # Calculate Stochastic Oscillator
            low_min = data['Low'].rolling(window=14).min()
            high_max = data['High'].rolling(window=14).max()
            data['%K'] = (data['Close'] - low_min) * 100 / (high_max - low_min)
            data['%D'] = data['%K'].rolling(window=3).mean()
            data['Slow %D'] = data['%D'].rolling(window=3).mean()

            # Calculate Bollinger Bands
            data['SMA_20'] = data['Close'].rolling(window=20).mean()
            data['STD_20'] = data['Close'].rolling(window=20).std()
            data['Upper_BB'] = data['SMA_20'] + (data['STD_20'] * 2)
            data['Lower_BB'] = data['SMA_20'] - (data['STD_20'] * 2)

            # Calculate Buy/Sell Signals for Bollinger Bands
            data['BB_Buy_Signal'] = (data['Close'] <= data['Lower_BB']) & (data['Close'].shift(1) > data['Lower_BB'].shift(1))
            data['BB_Sell_Signal'] = (data['Close'] >= data['Upper_BB']) & (data['Close'].shift(1) < data['Upper_BB'].shift(1))

            # Calculate Ichimoku Cloud
            high_9 = data['High'].rolling(window=9).max()
            low_9 = data['Low'].rolling(window=9).min()
            data['Tenkan_Sen'] = (high_9 + low_9) / 2
            high_26 = data['High'].rolling(window=26).max()
            low_26 = data['Low'].rolling(window=26).min()
            data['Kijun_Sen'] = (high_26 + low_26) / 2
            data['Senkou_Span_A'] = ((data['Tenkan_Sen'] + data['Kijun_Sen']) / 2).shift(26)
            high_52 = data['High'].rolling(window=52).max()
            low_52 = data['Low'].rolling(window=52).min()
            data['Senkou_Span_B'] = ((high_52 + low_52) / 2).shift(26)
            data['Chikou_Span'] = data['Close'].shift(-26)

            # Calculate Ichimoku Cloud Buy/Sell Signals
            data['Ichimoku_Buy_Signal'] = (data['Close'] > data['Senkou_Span_A']) & (data['Close'] > data['Senkou_Span_B']) & (data['Close'] > data['Kijun_Sen']) & (data['Close'].shift(1) <= data['Kijun_Sen'].shift(1))
            data['Ichimoku_Sell_Signal'] = (data['Close'] < data['Senkou_Span_A']) & (data['Close'] < data['Senkou_Span_B']) & (data['Close'] < data['Kijun_Sen']) & (data['Close'].shift(1) >= data['Kijun_Sen'].shift(1))

            # Calculate Long Lower Tail (LLT)
            data['Body'] = abs(data['Close'] - data['Open'])
            data['Lower_Shadow'] = data[['Open', 'Close']].min(axis=1) - data['Low']
            data['Upper_Shadow'] = data['High'] - data[['Open', 'Close']].max(axis=1)
            
            data['Hammer'] = (
                (data['Lower_Shadow'] > 2 * data['Body']) &  # Long lower shadow
                (data['Upper_Shadow'] <= 0.1 * data['Body']) &  # Little to no upper shadow
                (data['Body'] > 0)  # Ensure it's not a doji
            )

            # Calculate Engulfing Patterns
            data['Bullish_Engulfing'] = (data['Open'].shift(1) > data['Close'].shift(1)) & \
                                        (data['Close'] > data['Open'].shift(1)) & \
                                        (data['Open'] < data['Close'].shift(1))
            data['Bearish_Engulfing'] = (data['Open'].shift(1) < data['Close'].shift(1)) & \
                                        (data['Close'] < data['Open'].shift(1)) & \
                                        (data['Open'] > data['Close'].shift(1))

            # Calculate Marubozu Candles
            body_size = abs(data['Close'] - data['Open'])
            candle_size = data['High'] - data['Low']
            data['Marubozu'] = (body_size / candle_size) > 0.95

            # Calculate RSI
            delta = data['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            data['RSI'] = 100 - (100 / (1 + rs))

            # Calculate ADX
            high = data['High']
            low = data['Low']
            close = data['Close']
            period = 14

            plus_dm = high.diff()
            minus_dm = low.diff()
            plus_dm[plus_dm < 0] = 0
            minus_dm[minus_dm > 0] = 0

            tr1 = pd.DataFrame(high - low)
            tr2 = pd.DataFrame(abs(high - close.shift(1)))
            tr3 = pd.DataFrame(abs(low - close.shift(1)))
            frames = [tr1, tr2, tr3]
            tr = pd.concat(frames, axis=1, join='inner').max(axis=1)
            atr = tr.rolling(period).mean()

            plus_di = 100 * (plus_dm.rolling(period).mean() / atr)
            minus_di = abs(100 * (minus_dm.rolling(period).mean() / atr))
            dx = (abs(plus_di - minus_di) / abs(plus_di + minus_di)) * 100
            adx = ((dx.shift(1) * (period - 1)) + dx) / period
            adx_smooth = adx.rolling(period).mean()

            data['ADX'] = adx_smooth
            data['Plus_DI'] = plus_di
            data['Minus_DI'] = minus_di

            return data
        except Exception as e:
            self.logger.error(f"Error calculating indicators: {str(e)}")
            self.logger.error(traceback.format_exc())
            raise


class StockManager(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("Manage Saved Stocks")
        self.geometry("600x600")
        self.minsize(400, 400)
        self.stocks = self.load_stocks()
        self.filtered_stocks = self.stocks.copy()  # New attribute for filtered stocks
        self.selected_symbol = None
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_listbox)
        
        self.create_widgets()
        self.update_listbox()
        
        self.transient(parent)
        self.grab_set()

    def create_widgets(self):
        # Main frame
        main_frame = ctk.CTkFrame(self)
        main_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(1, weight=1)

        # Search frame
        search_frame = ctk.CTkFrame(main_frame)
        search_frame.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        search_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(search_frame, text="Search:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, width=200)
        search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Listbox for stocks
        self.stock_listbox = tk.Listbox(main_frame)
        self.stock_listbox.grid(row=1, column=0, sticky="nsew")

        # Scrollbar for listbox
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=self.stock_listbox.yview)
        scrollbar.grid(row=1, column=1, sticky="ns")
        self.stock_listbox.config(yscrollcommand=scrollbar.set)

        # Button frame
        button_frame = ctk.CTkFrame(self)
        button_frame.grid(row=1, column=0, pady=10, sticky="ew")
        button_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        # Buttons
        ctk.CTkButton(button_frame, text="Add", command=self.add_stock).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Edit", command=self.edit_stock).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Delete", command=self.delete_stock).grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Select", command=self.select_stock).grid(row=0, column=3, padx=5, pady=5, sticky="ew")



    def load_stocks(self):
        try:
            with open('saved_stocks.json', 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            return []

    def save_stocks(self):
        with open('saved_stocks.json', 'w') as f:
            json.dump(self.stocks, f, indent=2)


    def update_listbox(self, *args):
        self.stock_listbox.delete(0, tk.END)
        search_term = self.search_var.get().lower()
        self.filtered_stocks = []  # Reset filtered stocks
        for stock in self.stocks:
            if search_term in stock['symbol'].lower() or search_term in stock['name'].lower():
                self.stock_listbox.insert(tk.END, f"{stock['symbol']} - {stock['name']}")
                self.filtered_stocks.append(stock)  # Add to filtered stocks

    def update_tree(self, *args):
        self.tree.delete(*self.tree.get_children())
        search_term = self.search_var.get().lower()

        for stock in self.stock_viewer.saved_stocks:
            symbol = stock['symbol']
            name = stock['name']
            
            if search_term in symbol.lower() or search_term in name.lower():
                self.tree.insert('', 'end', values=(symbol, name))

    def sort_treeview(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)

        self.tree.heading(col, command=lambda: self.sort_treeview(col, not reverse))



    def add_stock(self):
        dialog = StockDialog(self, title="Add Stock")
        self.wait_window(dialog)
        if dialog.result:
            self.stocks.append(dialog.result)
            self.save_stocks()
            self.update_listbox()

    def edit_stock(self):
        selected = self.stock_listbox.curselection()
        if selected:
            index = selected[0]
            stock = self.filtered_stocks[index]  # Use filtered_stocks instead of stocks
            dialog = StockDialog(self, title="Edit Stock", initial_symbol=stock['symbol'], initial_name=stock['name'])
            self.wait_window(dialog)
            if dialog.result:
                # Update the stock in both filtered_stocks and stocks
                for i, s in enumerate(self.stocks):
                    if s['symbol'] == stock['symbol']:
                        self.stocks[i] = dialog.result
                        break
                self.filtered_stocks[index] = dialog.result
                self.save_stocks()
                self.update_listbox()
        else:
            CTkMessagebox(title="No Selection", message="Please select a stock to edit.", icon="warning")




    def delete_stock(self):
        selected = self.stock_listbox.curselection()
        if selected:
            index = selected[0]
            stock = self.filtered_stocks[index]  # Use filtered_stocks instead of stocks
            confirm = CTkMessagebox(title="Confirm Deletion", 
                                    message=f"Are you sure you want to delete {stock['symbol']} - {stock['name']}?",
                                    icon="question", option_1="No", option_2="Yes")
            if confirm.get() == "Yes":
                self.stocks.remove(stock)  # Remove from main stocks list
                self.filtered_stocks.pop(index)  # Remove from filtered list
                self.save_stocks()
                self.update_listbox()
        else:
            CTkMessagebox(title="No Selection", message="Please select a stock to delete.", icon="warning")


    def select_stock(self):
        selected = self.stock_listbox.curselection()
        if selected:
            index = selected[0]
            stock = self.filtered_stocks[index]  # Use filtered_stocks instead of stocks
            self.selected_symbol = stock['symbol']
            self.destroy()
        else:
            CTkMessagebox(title="No Selection", message="Please select a stock.", icon="warning")

    def refresh_tree(self):
        # Save the current scroll position
        current_selection = self.tree.selection()
        current_scroll = self.tree.yview()

        # Clear the existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Repopulate with current data
        for stock in self.stock_viewer.saved_stocks:
            self.tree.insert('', 'end', values=(stock['symbol'], stock['name']))

        # Restore the scroll position
        self.tree.yview_moveto(current_scroll[0])

        # Restore the selection if possible
        if current_selection:
            for item in self.tree.get_children():
                if self.tree.item(item)['values'][0] == self.tree.item(current_selection[0])['values'][0]:
                    self.tree.selection_set(item)
                    self.tree.see(item)
                    break
                    
                    
    def handle_selected_stock(self, symbol):
        self.stock_entry.delete(0, tk.END)
        self.stock_entry.insert(0, symbol)

class StockDialog(ctk.CTkToplevel):
    def __init__(self, parent, title, initial_symbol="", initial_name=""):
        super().__init__(parent)
        self.title(title)
        self.geometry("320x200")  # Increased width
        self.minsize(320, 200)
        self.result = None

        ctk.CTkLabel(self, text="Stock Symbol:").pack(pady=1)
        self.symbol_entry = ctk.CTkEntry(self, width=300)  # Increased width
        self.symbol_entry.pack(pady=5)
        self.symbol_entry.insert(0, initial_symbol)

        ctk.CTkLabel(self, text="Company Name:").pack(pady=1)
        self.name_entry = ctk.CTkEntry(self, width=300)  # Increased width
        self.name_entry.pack(pady=5)
        self.name_entry.insert(0, initial_name)

        ctk.CTkButton(self, text="Save", command=self.save).pack(pady=5)

        self.grab_set()

    def save(self):
        symbol = self.symbol_entry.get().strip().upper()
        name = self.name_entry.get().strip()
        if symbol and name:
            self.result = {"symbol": symbol, "name": name}
            self.destroy()
        else:
            CTkMessagebox(title="Invalid Input", message="Both symbol and name are required.", icon="warning")



class SavedStocksManager(ctk.CTkToplevel):
    def __init__(self, parent, stock_viewer):
        super().__init__(parent)
        self.stock_viewer = stock_viewer
        self.title("Manage Saved Stocks")
        self.geometry("800x400")
        
        self.grab_set()
        self.transient(parent)
        self.focus_force()
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.update_tree)

        self.create_widgets()
        self.update_tree()

    def create_widgets(self):
        # Search frame
        search_frame = ctk.CTkFrame(self)
        search_frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")
        search_frame.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(search_frame, text="Search:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var, width=200)
        search_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        # Treeview
        self.tree = ttk.Treeview(self, columns=('Symbol', 'Name'), show='headings')
        self.tree.heading('Symbol', text='Stock Symbol', command=lambda: self.sort_treeview('Symbol', False))
        self.tree.heading('Name', text='Company Name', command=lambda: self.sort_treeview('Name', False))
        self.tree.column('Symbol', width=100, minwidth=50)
        self.tree.column('Name', width=300, minwidth=150)
        self.tree.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

        # Scrollbar
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.tree.yview)
        scrollbar.grid(row=2, column=1, sticky="ns")
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Buttons
        button_frame = ctk.CTkFrame(self)
        button_frame.grid(row=3, column=0, pady=10, sticky="ew")
        button_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        ctk.CTkButton(button_frame, text="Add", command=self.add_stock).grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Edit", command=self.edit_stock).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Delete", command=self.delete_stock).grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        ctk.CTkButton(button_frame, text="Select", command=self.select_stock).grid(row=0, column=3, padx=5, pady=5, sticky="ew")



    def update_tree(self, *args):
        self.tree.delete(*self.tree.get_children())
        search_term = self.search_var.get().lower()

        for stock in self.stock_viewer.saved_stocks:
            symbol = stock['symbol']
            name = stock['name']
            
            if search_term in symbol.lower() or search_term in name.lower():
                self.tree.insert('', 'end', values=(symbol, name))


    def sort_treeview(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)

        self.tree.heading(col, command=lambda: self.sort_treeview(col, not reverse))

    def add_stock(self):
        dialog = AddEditStockDialog(self, self.stock_viewer, mode='add')
        self.wait_window(dialog)
        self.update_tree()


    def edit_stock(self):
        selected = self.tree.selection()
        if selected:
            item = self.tree.item(selected[0])
            symbol, name = item['values']
            dialog = AddEditStockDialog(self, self.stock_viewer, mode='edit', initial_symbol=symbol, initial_name=name)
            self.wait_window(dialog)
            self.update_tree()
        else:
            CTkMessagebox(title="No Selection", message="Please select a stock to edit.", icon="warning")


    def delete_stock(self):
        selected = self.tree.selection()
        if selected:
            item = self.tree.item(selected[0])
            symbol, name = item['values']
            confirm = CTkMessagebox(title="Confirm Deletion", 
                                    message=f"Are you sure you want to delete {name} ({symbol})?",
                                    icon="question", option_1="Yes", option_2="No")
            if confirm.get() == "Yes":
                # Remove from treeview
                self.tree.delete(selected[0])
                
                # Remove from saved_stocks list
                self.stock_viewer.saved_stocks = [stock for stock in self.stock_viewer.saved_stocks if stock['symbol'] != symbol]
                
                # Save changes
                self.stock_viewer.save_stocks()
                
                # Log the deletion
                print(f"Deleted stock: {symbol} ({name})")
                print(f"Updated saved_stocks: {self.stock_viewer.saved_stocks}")
                
                CTkMessagebox(title="Deletion Successful", message=f"Stock '{name}' has been deleted.", icon="info")
                
                # Refresh the treeview
                self.update_tree()
        else:
            CTkMessagebox(title="No Selection", message="Please select a stock to delete.", icon="warning")

                          
    def select_stock(self):
        selected = self.tree.selection()
        if selected:
            item = self.tree.item(selected[0])
            symbol, _ = item['values']
            self.stock_viewer.stock_entry.delete(0, tk.END)
            self.stock_viewer.stock_entry.insert(0, symbol)  # Use the full symbol, including leading zeros
            self.destroy()
        else:
            CTkMessagebox(title="No Selection", message="Please select a stock.", icon="warning")

    def refresh_tree(self):
        # Clear the existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Repopulate with current data
        for stock in self.stock_viewer.saved_stocks:
            self.tree.insert('', 'end', values=(stock['symbol'], stock['name']))

class AddEditStockDialog(ctk.CTkToplevel):
    def __init__(self, parent, stock_viewer, mode='add', initial_symbol='', initial_name=''):
        super().__init__(parent)
        self.parent = parent
        self.stock_viewer = stock_viewer
        self.mode = mode
        self.title("Add Stock" if mode == 'add' else "Edit Stock")
        self.geometry("400x200")

        self.grab_set()
        self.transient(parent)
        self.focus_set()
  #      self.attributes('-topmost', True)

        self.create_widgets(initial_symbol, initial_name)

    def create_widgets(self, initial_symbol, initial_name):
        frame = ctk.CTkFrame(self)
        frame.pack(padx=20, pady=20, fill=ctk.BOTH, expand=True)

        ctk.CTkLabel(frame, text="Stock Symbol:").grid(row=0, column=0, padx=5, pady=10, sticky='e')
        self.symbol_entry = ctk.CTkEntry(frame, width=200)
        self.symbol_entry.grid(row=0, column=1, padx=5, pady=10, sticky='ew')
        self.symbol_entry.insert(0, initial_symbol)

        ctk.CTkLabel(frame, text="Company Name:").grid(row=1, column=0, padx=5, pady=10, sticky='e')
        self.name_entry = ctk.CTkEntry(frame, width=200)
        self.name_entry.grid(row=1, column=1, padx=5, pady=10, sticky='ew')
        self.name_entry.insert(0, initial_name)

        button_frame = ctk.CTkFrame(frame)
        button_frame.grid(row=2, column=0, columnspan=2, pady=20)

        ctk.CTkButton(button_frame, text="Save", command=self.save_stock).pack(side=ctk.LEFT, padx=10)
        ctk.CTkButton(button_frame, text="Cancel", command=self.destroy).pack(side=ctk.LEFT, padx=10)

    def save_stock(self):
        symbol = self.symbol_entry.get().strip().upper()
        name = self.name_entry.get().strip()

        if not symbol or not name:
            CTkMessagebox(title="Invalid Input", message="Both symbol and name are required.", icon="warning")
            return

        if self.mode == 'add':
            self.stock_viewer.saved_stocks.append({"symbol": symbol, "name": name})
        else:
            for stock in self.stock_viewer.saved_stocks:
                if stock['symbol'] == symbol:
                    stock['name'] = name
                    break

        self.stock_viewer.save_stocks()
        self.destroy()




class NewsApp:
    def __init__(self, master):
        self.master = master
        self.master.title("CatalystWatcher - v1.0.5")
        self.master.geometry("1200x700")  # Set initial size
        self.master.minsize(800, 600)  # Set minimum size
        
        # Initialize PyQt5 application
        self.qt_app = QApplication.instance()
        if not self.qt_app:
            self.qt_app = QApplication(sys.argv)

        self.filter_var = tk.StringVar()
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()
        self.time_var = tk.StringVar()

        self.sentiment_filter = None

        self.news_data = []
        self.filtered_data = []
        self.selected_feeds = []

        self.progress_queue = queue.Queue()

        self.load_icons()

        self.load_rss_feeds()
        self.load_keywords()

        self.columns = {
            "#": 40,
            "Sentiment": 70,
            "Date": 100,
            "Time": 100,
            "Title": 600,
            "Description": 1800,
            "Source": 200,
            "Website": 200
        }
        

        self.create_widgets()
        self.create_context_menu()

        # Initialize TradingView widget
        self.tradingview_widget = None

        # Schedule opening TradingView widget after the main window is fully initialized
        self.master.after(100, self.open_tradingview_widget)

        self.economic_calendar_widget = None
        self.trading_chart_widget = None
        self.bursa_heatmap_widget = None
        self.us_heatmap_widget = None
        self.fear_greed_widget = None

        self.filter_var.trace("w", self.apply_filter)
        self.start_date_var.trace("w", self.apply_filter)
        self.end_date_var.trace("w", self.apply_filter)
        self.time_var.trace("w", self.apply_filter)

        self.fetch_thread = None
        self.is_running = True

        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        self.master.update_idletasks()  # Ensure the window size is updated
        self.center_window()

    def load_icons(self):
        self.icons = {
            'us_heatmap': QIcon(os.path.join('icons', 'us_heatmap_icon.png')),
            'bursa_heatmap': QIcon(os.path.join('icons', 'bursa_heatmap_icon.png')),
            'economic_calendar': QIcon(os.path.join('icons', 'economic_calendar_icon.png')),
            'trading_chart': QIcon(os.path.join('icons', 'trading_chart_icon.png')),
            'fear_greed': QIcon(os.path.join('icons', 'fear_greed_icon.png')),
            'tradingview': QIcon(os.path.join('icons', 'tradingview_icon.png')),
        }

    def center_window(self):
        # Get the configured window size
        width = 1200  # Use the initial width you set
        height = 700  # Use the initial height you set

        # Calculate position
        screen_width = self.master.winfo_screenwidth()
        screen_height = self.master.winfo_screenheight()

        x = (screen_width - width) // 2
        y = (screen_height - height) // 2

        # Set the position of the window to the center of the screen
        self.master.geometry(f'{width}x{height}+{x}+{y}')
        
    def open_stock_price_viewer(self):
        StockPriceViewer(self.master)

    def open_tradingview_widget(self):
        if not self.tradingview_widget:
            self.tradingview_widget = run_tradingview_widget()
        else:
            self.tradingview_widget.show()
            self.tradingview_widget.center_top()

    def open_economic_calendar_widget(self):
        if not self.economic_calendar_widget:
            self.economic_calendar_widget = run_economic_calendar_widget(self.icons['economic_calendar'])
        else:
            self.economic_calendar_widget.show()
            self.economic_calendar_widget.center_top()

    def open_trading_chart_widget(self):
        if not self.trading_chart_widget:
            self.trading_chart_widget = run_trading_chart_widget()
        else:
            self.trading_chart_widget.show()
            self.trading_chart_widget.center_top()

    # Update other widget opening methods similarly
    def open_bursa_heatmap_widget(self):
        self.bursa_heatmap_widget = run_bursa_heatmap_widget(self.icons['bursa_heatmap'])
        self.bursa_heatmap_widget.show()
        self.bursa_heatmap_widget.center_top()


    def open_us_heatmap_widget(self):
        if not self.us_heatmap_widget:
            self.us_heatmap_widget = run_us_heatmap_widget(self.icons['us_heatmap'])
        else:
            self.us_heatmap_widget.show()
            self.us_heatmap_widget.center()

    # Add this method to your NewsApp class
    def open_ai_operation(self, choice):
        if choice == "AI Summary":
            create_ai_summary_widget(self.master, self.filtered_data)
        elif choice == "AI Analysis":
            self.run_ai_analysis()
            
    def run_ai_analysis(self):
        try:
            # Get the directory of the current script
            current_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Construct the path to app.py
            app_path = os.path.join(current_dir, 'app.py')
            
            # Run Streamlit as a separate process
            subprocess.Popen([sys.executable, '-m', 'streamlit', 'run', app_path])
            
            # Open the default web browser to the Streamlit app
            webbrowser.open('http://localhost:8501')
            
            CTkMessagebox(title="AI Analysis", message="AI Analysis app is running in your web browser.", icon="info")
        except Exception as e:
            CTkMessagebox(title="Error", message=f"An error occurred while starting AI Analysis: {str(e)}", icon="cancel")



#    def open_fear_greed_widget(self, choice):
#        market = "MY" if "MY" in choice else "US"
#        if self.fear_greed_widget:
#            self.fear_greed_widget.close()
#        self.fear_greed_widget = run_fear_greed_widget(market)
#        self.fear_greed_widget.show()


    def open_fear_greed_widget(self, choice):
        if "US" in choice:
            if self.fear_greed_widget:
                self.fear_greed_widget.close()
            self.fear_greed_widget = run_us_fear_greed_widget()
        else:
            market = "MY"
            if self.fear_greed_widget:
                self.fear_greed_widget.close()
            self.fear_greed_widget = run_fear_greed_widget(market)
        self.fear_greed_widget.show()
      

    def open_active_gainers_losers(self):
        # Create and start a new process for the StocksWidget
        process = multiprocessing.Process(target=run_stocks_widget)
        process.start()

    
    def create_widgets(self):
        self.master.grid_columnconfigure(0, weight=1)
        self.master.grid_rowconfigure(3, weight=1)

        # Input frame
        input_frame = ctk.CTkFrame(self.master)
        input_frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")
        input_frame.grid_columnconfigure(7, weight=1)

        ctk.CTkLabel(input_frame, text="Filter:").grid(row=0, column=0, padx=5, pady=5)
        self.filter_entry = ctk.CTkEntry(input_frame, width=150, textvariable=self.filter_var)
        self.filter_entry.grid(row=0, column=1, padx=5, pady=5)

        ctk.CTkLabel(input_frame, text="From:").grid(row=0, column=2, padx=5, pady=5)
        self.start_date_entry = DateEntry(input_frame, width=10, background='darkblue', foreground='white', borderwidth=2, textvariable=self.start_date_var)
        self.start_date_entry.grid(row=0, column=3, padx=5, pady=5)

        ctk.CTkLabel(input_frame, text="To:").grid(row=0, column=4, padx=5, pady=5)
        self.end_date_entry = DateEntry(input_frame, width=10, background='darkblue', foreground='white', borderwidth=2, textvariable=self.end_date_var)
        self.end_date_entry.grid(row=0, column=5, padx=5, pady=5)

        ctk.CTkLabel(input_frame, text="Time (HH:MM):").grid(row=0, column=6, padx=5, pady=5)
        self.time_entry = ctk.CTkEntry(input_frame, width=80, textvariable=self.time_var)
        self.time_entry.grid(row=0, column=7, padx=5, pady=5, sticky="w")

        # Button frame
        button_frame = ctk.CTkFrame(self.master)
        button_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        button_frame.grid_columnconfigure((0, 1, 2, 3, 4, 5, 6, 7, 8, 9), weight=1)

        # First row of buttons (default green color)
        ctk.CTkButton(button_frame, text="Fetch News", command=self.fetch_news_thread).grid(row=0, column=0, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Positive News", command=lambda: self.filter_sentiment("positive")).grid(row=0, column=1, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Negative News", command=lambda: self.filter_sentiment("negative")).grid(row=0, column=2, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="All News", command=self.reset_sentiment_filter).grid(row=0, column=3, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Export", command=self.export_to_excel).grid(row=0, column=4, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Manage Feeds", command=self.manage_rss_feeds).grid(row=0, column=5, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Manage Keywords", command=self.manage_keywords).grid(row=0, column=6, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Select Feeds", command=self.select_feeds).grid(row=0, column=7, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Stock Prices", command=self.open_stock_price_viewer).grid(row=0, column=8, padx=2, pady=5)

        # CustomTkinter blue dark theme style for second row buttons
        blue_button_style = {
            "fg_color": "#1F538D",  # Dark blue
            "hover_color": "#14375E",  # Darker blue for hover
            "text_color": "white"
        }

        # CustomTkinter Orange dark theme style for second row buttons
        orange_button_style = {
            "fg_color": "#ff9933", 
            "hover_color": "#cc6600",  # Dark for hover
            "text_color": "white"
        }

        # Second row of buttons (blue dark theme)
        ctk.CTkButton(button_frame, text="Ticker", command=self.open_tradingview_widget, **blue_button_style).grid(row=1, column=0, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Economic Calendar", command=self.open_economic_calendar_widget, **blue_button_style).grid(row=1, column=1, padx=2, pady=5)
        ctk.CTkButton(button_frame, text="Trading Chart", command=self.open_trading_chart_widget, **blue_button_style).grid(row=1, column=2, padx=2, pady=5)
        
        # Heatmap dropdown (blue dark theme)
        self.heatmap_var = ctk.StringVar(value="         Heatmap")
        heatmap_dropdown = ctk.CTkOptionMenu(
            button_frame,
            values=["MY Heatmap", "US Heatmap"],
            command=self.open_selected_heatmap,
            variable=self.heatmap_var,
            fg_color="#1F538D",
            button_color="#1F538D",
            button_hover_color="#14375E",
            text_color="white"
        )
        heatmap_dropdown.grid(row=1, column=3, padx=2, pady=5)

  
        self.ai_var = ctk.StringVar(value="    AI Operations")
        ai_dropdown = ctk.CTkOptionMenu(
            button_frame,
            values=["AI Summary", "AI Analysis"],
            command=self.open_ai_operation,
            variable=self.ai_var,
            fg_color="#ff9933",  # Orange color
            button_color="#ff9933",
            button_hover_color="#cc6600",
            text_color="white"
        )
        ai_dropdown.grid(row=1, column=4, padx=2, pady=5)
  
  
        ctk.CTkButton(button_frame, text="Fear & Greed", command=self.open_fear_greed_widget, **blue_button_style).grid(row=1, column=5, padx=2, pady=5)

        # Add the new "Active Gainers/Losers" button
        ctk.CTkButton(button_frame, text="Gainers/Losers", command=self.open_active_gainers_losers, **blue_button_style).grid(row=1, column=6, padx=2, pady=5)


        # Replace the existing Fear and Greed button with a dropdown
        self.fear_greed_var = ctk.StringVar(value="Fear & Greed")
        fear_greed_dropdown = ctk.CTkOptionMenu(
            button_frame,
            values=["MY Fear & Greed", "US Fear & Greed"],
            command=self.open_fear_greed_widget,
            variable=self.fear_greed_var,
            fg_color="#1F538D",
            button_color="#1F538D",
            button_hover_color="#14375E",
            text_color="white"
        )
        fear_greed_dropdown.grid(row=1, column=5, padx=2, pady=5)

        # News count label
        self.news_count_label = ctk.CTkLabel(self.master, text="Total News: 0")
        self.news_count_label.grid(row=2, column=0, padx=10, pady=5, sticky="e")

        # Results frame
        results_frame = ctk.CTkFrame(self.master)
        results_frame.grid(row=3, column=0, padx=10, pady=5, sticky="nsew")
        results_frame.grid_columnconfigure(0, weight=1)
        results_frame.grid_rowconfigure(0, weight=1)

        # Treeview
        self.tree = ttk.Treeview(results_frame, columns=list(self.columns.keys()), show="headings")
        
        for col, width in self.columns.items():
            self.tree.heading(col, text=col, command=lambda _col=col: self.sort_treeview(_col, False))
            self.tree.column(col, width=width, minwidth=50)

        self.tree.grid(row=0, column=0, sticky="nsew")
        self.tree.bind("<Button-1>", self.on_treeview_click)

        # Scrollbars
        vsb = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # Copyright label
        copyright_label = ctk.CTkLabel(self.master, text="© Created by Hariadi Nawawi | 2024")
        copyright_label.grid(row=4, column=0, padx=10, pady=5, sticky="e")
    def open_stock_price_viewer(self):
        StockPriceViewer(self.master)

    def open_selected_heatmap(self, choice):
        if choice == "MY Heatmap":
            self.open_bursa_heatmap_widget()
        elif choice == "US Heatmap":
            self.open_us_heatmap_widget()

    def on_column_resize(self, event):
        if self.tree.identify_region(event.x, event.y) == "separator":
            column = self.tree.identify_column(event.x)
            if column:
                col_name = self.tree["columns"][int(column[1:]) - 1]
                new_width = self.tree.column(column, "width")
                self.column_widths[col_name] = new_width
                self.save_column_widths()
                logger.debug(f"Column {col_name} resized to {new_width}")
                

    def load_column_widths(self):
        try:
            with open('column_widths.json', 'r') as f:
                widths = json.load(f)
            # Ensure all columns have a width
            for col in self.columns:
                if col not in widths:
                    widths[col] = 100  # Default width
            logger.debug(f"Loaded column widths: {widths}")
            return widths
        except FileNotFoundError:
            logger.debug("Column widths file not found, using default widths")
            return {col: 100 for col in self.columns}  # Default widths

    def save_column_widths(self):
        with open('column_widths.json', 'w') as f:
            json.dump(self.column_widths, f)
        logger.debug(f"Saved column widths: {self.column_widths}")

    def apply_column_widths(self):
        for col in self.tree["columns"]:
            width = self.column_widths.get(col, 100)
            self.tree.column(col, width=width, minwidth=50)
        logger.debug("Applied column widths")


    def on_closing(self):
        # Close the TradingView widget if it's open
        if self.tradingview_widget:
            self.tradingview_widget.close()
        
        # Close the Economic Calendar widget if it's open
        if self.economic_calendar_widget:
            self.economic_calendar_widget.close()
        
        # Close the Trading Chart widget if it's open
        if self.trading_chart_widget:
            self.trading_chart_widget.close()
        
        # Close the Bursa Malaysia Heatmap widget if it's open
        if self.bursa_heatmap_widget:
            self.bursa_heatmap_widget.close()
        
        # Close the US Heatmap widget if it's open
        if self.us_heatmap_widget:
            self.us_heatmap_widget.close()
        
        # Close the Fear and Greed widget if it's open
        if self.fear_greed_widget:
            self.fear_greed_widget.close()

        # Close the Active Gainers/Losers widget if it's open
        if hasattr(self, 'active_gainers_losers_widget') and self.active_gainers_losers_widget:
            self.active_gainers_losers_widget.close()

        # Stop any ongoing threads
        for thread in threading.enumerate():
            if thread != threading.main_thread():
                thread.join(timeout=1)
        self.master.destroy()

    def fetch_news_thread(self):
        if not self.selected_feeds:
            messagebox.showwarning("No Feeds Selected", "Please select feeds before fetching news.")
            return
        self.progress_bar = ProgressBar(self.master)
        self.progress_bar.start()
        self.fetch_thread = threading.Thread(target=self.fetch_news, daemon=True)
        self.fetch_thread.start()
        self.process_progress_queue()

    def strip_html_tags(self, html_text):
        # Remove all HTML tags
        text = re.sub('<[^<]+?>', '', html_text)
        # Unescape HTML entities
        text = unescape(text)
        # Remove extra whitespace
        text = ' '.join(text.split())
        return text


       
    def process_progress_queue(self):
        try:
            while True:
                message = self.progress_queue.get_nowait()
                if message[0] == "status":
                    self.progress_bar.update_status(message[1])
                elif message[0] == "feed":
                    self.progress_bar.update_feed(message[1])
                elif message[0] == "item":
                    self.progress_bar.update_item_count(message[1], message[2])
                self.progress_queue.task_done()
        except queue.Empty:
            self.master.after(100, self.process_progress_queue)


    def fetch_news(self):
        try:
            self.news_data = []
            self.master.after(0, self.update_treeview, [])

            self.progress_queue.put(("status", "Fetching latest financial news..."))
            logging.debug("Fetching latest financial news")

            feed_stats = []
            total_feeds = len(self.selected_feeds)
            for feed_index, feed in enumerate(self.selected_feeds, 1):
                if not self.progress_bar.is_alive or not self.is_running:
                    logging.info("News fetch cancelled.")
                    return

                try:
                    if isinstance(feed, dict):
                        feed_name = feed.get('name', 'Unknown')
                        feed_url = feed.get('url', '')
                        website = feed.get('website', '')
                    elif isinstance(feed, (list, tuple)) and len(feed) >= 2:
                        feed_name, feed_url = feed[:2]
                        website = feed[2] if len(feed) > 2 else ''
                    else:
                        logging.error(f"Invalid feed structure: {feed}")
                        continue

                    if not feed_url:
                        logging.error(f"No URL provided for feed: {feed_name}")
                        continue

                    self.progress_queue.put(("status", f"Processing feed {feed_index} of {total_feeds}"))
                    self.progress_queue.put(("feed", feed_name))

                    time.sleep(2)  # Add a delay between requests to avoid rate limiting
                    
                    response = requests.get(feed_url, timeout=10)
                    response.raise_for_status()
                    parsed_feed = feedparser.parse(response.content)
                    source = parsed_feed.feed.title if 'title' in parsed_feed.feed else feed_name
                    items_count = 0

                    logging.debug(f"Processing feed: {source} ({feed_url})")
                    logging.debug(f"Number of entries: {len(parsed_feed.entries)}")

                    for entry_index, entry in enumerate(parsed_feed.entries, 1):
                        if not self.progress_bar.is_alive or not self.is_running:
                            logging.info("News fetch cancelled.")
                            return

                        self.progress_queue.put(("item", entry_index, len(parsed_feed.entries)))

                        title = entry.title
                        description = entry.get('description', '')
                        description = self.strip_html_tags(description)
                        
                        try:
                            # Use dateutil.parser to parse the date
                            date = parser.parse(entry.published)
                            if date.tzinfo is None:
                                date = date.replace(tzinfo=timezone.utc)
                        except (AttributeError, ValueError) as e:
                            logging.warning(f"Could not parse date for entry: {title}. Error: {str(e)}")
                            continue

                        sentiment = self.get_sentiment(title + " " + description)
                        
                        # Use datetime.now(timezone.utc) for comparison
                        if datetime.now(timezone.utc) - date <= timedelta(days=3):
                            item = {
                                'title': title,
                                'date': date.strftime('%Y-%m-%d'),
                                'time': date.strftime('%H:%M:%S'),
                                'sentiment': sentiment,
                                'description': description,
                                'source': source,
                                'website': website
                            }
                            self.news_data.append(item)
                            items_count += 1
                        else:
                            logging.debug(f"Skipping old news: {title} ({date.isoformat()})")

                    feed_stats.append(f"{source} ({feed_url}): {items_count} items")
                    logging.debug(f"Processed {items_count} items from {source}")
                except requests.exceptions.RequestException as e:
                    error_message = str(e)
                    if "429" in error_message:
                        error_message = "Rate limit exceeded. Please try again later."
                    logging.error(f"Error fetching feed {feed_url}: {error_message}")
                    feed_stats.append(f"{feed_url}: Error - {error_message}")
                except Exception as e:
                    logging.error(f"Unexpected error processing feed {feed_name}: {str(e)}")
                    feed_stats.append(f"{feed_name}: Unexpected Error - {str(e)}")

            if self.progress_bar.is_alive and self.is_running:
                self.progress_queue.put(("status", "Sorting and updating news data..."))
                self.news_data.sort(key=lambda x: x['sentiment'], reverse=True)
                self.filtered_data = self.news_data.copy()
                self.master.after(0, self.update_treeview, self.filtered_data)
                self.master.after(0, self.show_feed_stats, feed_stats)
                logging.debug(f"Processed {len(self.news_data)} news items in total.")

        except Exception as e:
            logging.error(f"Error in fetch_news: {str(e)}")
            logging.error(traceback.format_exc())
            if self.is_running:
                error_message = f"An error occurred while fetching news: {str(e)}"
                self.master.after(0, lambda: CTkMessagebox(title="Error", message=error_message, icon="cancel"))
        finally:
            if hasattr(self, 'progress_bar') and self.progress_bar.is_alive:
                self.master.after(0, self.progress_bar.stop)
        
    def show_feed_stats(self, feed_stats):
        stats_window = tk.Toplevel(self.master)
        stats_window.title("RSS Feed Statistics")
        stats_window.geometry("700x400")
        
        text_widget = tk.Text(stats_window, wrap=tk.WORD, width=80, height=20)
        text_widget.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        for stat in feed_stats:
            text_widget.insert(tk.END, stat + "\n\n")
        
        text_widget.config(state=tk.DISABLED)

    def get_sentiment(self, text):
        analysis = TextBlob(text)
        
        # Enhanced sentiment analysis
        score = analysis.sentiment.polarity
        
        for word in self.positive_keywords:
            if word.lower() in text.lower():
                score += 0.1
        for word in self.negative_keywords:
            if word.lower() in text.lower():
                score -= 0.1
        
        return max(-1, min(1, score))  # Ensure score is between -1 and 1

    def apply_filter(self, *args):
        if not hasattr(self, 'tree'):
            return  # Exit if tree is not yet initialized

        filter_text = self.filter_var.get().lower()
        start_date = self.start_date_var.get()
        end_date = self.end_date_var.get()
        time_filter = self.time_var.get()

        self.filtered_data = []
        for item in self.news_data:
            if filter_text in item['title'].lower() or filter_text in item['description'].lower():
                item_date = datetime.strptime(item['date'], '%Y-%m-%d').date()
                if start_date and end_date:
                    start = datetime.strptime(start_date, '%m/%d/%y').date()
                    end = datetime.strptime(end_date, '%m/%d/%y').date()
                    if not (start <= item_date <= end):
                        continue
                if time_filter and not item['time'].startswith(time_filter):
                    continue
                if self.sentiment_filter:
                    if self.sentiment_filter == "positive" and item['sentiment'] < 0.00:
                        continue
                    elif self.sentiment_filter == "negative" and item['sentiment'] > -0.01:
                        continue
                self.filtered_data.append(item)

        self.update_treeview(self.filtered_data)

    def filter_sentiment(self, sentiment):
        self.sentiment_filter = sentiment
        self.apply_filter()

    def reset_sentiment_filter(self):
        self.sentiment_filter = None
        self.filtered_data = self.news_data.copy()
        self.update_treeview(self.filtered_data)

    def sort_treeview(self, col, reverse):
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        l.sort(key=lambda t: float(t[0]) if col in ("#", "Sentiment") else t[0], reverse=reverse)

        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
            self.tree.set(k, "#", str(index + 1))  # Update the numbering

        self.tree.heading(col, command=lambda: self.sort_treeview(col, not reverse))

    def export_to_excel(self):
        if not self.filtered_data:
            messagebox.showwarning("Export Error", "No data to export.")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if file_path:
            try:
                df = pd.DataFrame(self.filtered_data)
                df = df[['sentiment', 'date', 'time', 'title', 'description', 'source', 'website']]  # Include 'website'
                
                df['sentiment'] = df['sentiment'].round(2)
                
                df.to_excel(file_path, index=False, engine='openpyxl')
                
                wb = load_workbook(file_path)
                ws = wb.active
                
                for idx, column in enumerate(df.columns, 1):
                    column_letter = get_column_letter(idx)
                    column_width = self.columns.get(column.capitalize(), 100)
                    ws.column_dimensions[column_letter].width = column_width / 7
                
                wb.save(file_path)
                
                result = messagebox.askyesno("Export Successful", f"Data exported to {file_path}\n\nDo you want to open the file?", icon='info')
                if result:
                    try:
                        if os.name == 'nt':  # For Windows
                            os.startfile(file_path)
                        elif os.name == 'posix':  # For macOS and Linux
                            subprocess.call(('open', file_path))
                    except Exception as e:
                        messagebox.showerror("Error", f"Unable to open the file: {str(e)}")
            except PermissionError:
                messagebox.showerror("Export Error", "Permission denied. The file may be in use or you may not have write permissions for this location. Please choose a different location or close any programs using the file.")
            except Exception as e:
                messagebox.showerror("Export Error", f"An error occurred while exporting: {str(e)}")




    def create_context_menu(self):
        self.context_menu = tk.Menu(self.master, tearoff=0)
        self.context_menu.add_command(label="Copy Title", command=self.copy_title)
        self.tree.bind("<Button-3>", self.show_context_menu)

    def show_context_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.context_menu.post(event.x_root, event.y_root)

    def copy_title(self):
        selected_items = self.tree.selection()
        if selected_items:
            selected_item = selected_items[0]
            title = self.tree.item(selected_item)['values'][4]  # Assuming title is at index 4
            pyperclip.copy(title)
            CTkMessagebox(title="Copied", message="Title copied to clipboard", icon="info")
        else:
            CTkMessagebox(title="No Selection", message="Please select a news item to copy its title.", icon="warning")



    def on_treeview_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            item = self.tree.identify_row(event.y)
            if item:
                values = self.tree.item(item)['values']
                # Assuming "Source" is the 7th column and "Website" is the 8th column
                if column == "#7":  # Source column
                    source = values[6]  # Index 6 corresponds to the 7th column
                    website = values[7]  # Website URL from the 8th column
                    if website:
                        webbrowser.open(website)
                    else:
                        CTkMessagebox(title="No Website", message=f"No website URL available for {source}", icon="info")
                elif column == "#8":  # Website column
                    website = values[7]  # Website URL from the 8th column
                    if website:
                        webbrowser.open(website)
                    else:
                        CTkMessagebox(title="No Website", message="No website URL available", icon="info")



    def update_treeview(self, news_data):
        self.tree.delete(*self.tree.get_children())
        for index, item in enumerate(news_data, start=1):
            self.tree.insert("", "end", values=(
                index, 
                f"{item['sentiment']:.2f}", 
                item['date'], 
                item['time'], 
                item['title'], 
                item['description'], 
                item['source'],
                item.get('website', '')  # Add the website URL to the treeview
            ))

        self.news_count_label.configure(text=f"Total News: {len(news_data)}")
                               
                    
                    

    def load_rss_feeds(self):
        try:
            with open('rss_feeds.json', 'r') as f:
                self.rss_feeds = json.load(f)
        except FileNotFoundError:
            self.rss_feeds = [
                {"name": "Yahoo Finance", "url": "https://finance.yahoo.com/news/rssindex"},
                {"name": "CNBC", "url": "https://www.cnbc.com/id/10000664/device/rss/rss.html"},
                {"name": "MarketWatch", "url": "https://feeds.content.dowjones.io/public/rss/mw_marketpulse"}
            ]
            self.save_rss_feeds()

    def save_rss_feeds(self):
        with open('rss_feeds.json', 'w') as f:
            json.dump(self.rss_feeds, f)

    def load_keywords(self):
        try:
            with open('keywords.json', 'r') as f:
                keywords = json.load(f)
                self.positive_keywords = keywords['positive']
                self.negative_keywords = keywords['negative']
        except FileNotFoundError:
            self.positive_keywords = [
                'surge', 'gain', 'rise', 'improve', 'growth', 'bull', 'rally', 'uptrend', 'outperform',
                'beat', 'exceed', 'positive', 'upgrade', 'strong', 'boom', 'recovery', 'breakthrough'
            ]
            self.negative_keywords = [
                'drop', 'fall', 'decline', 'loss', 'debt', 'bear', 'crash', 'downtrend', 'underperform',
                'miss', 'below', 'negative', 'downgrade', 'weak', 'bust', 'recession', 'setback'
            ]
            self.save_keywords()

    def save_keywords(self):
        keywords = {
            'positive': self.positive_keywords,
            'negative': self.negative_keywords
        }
        with open('keywords.json', 'w') as f:
            json.dump(keywords, f)
            
            
            
    def refresh_keywords(self):
        self.load_keywords()

    def manage_keywords(self):
        KeywordManager(self.master, self)
        self.recalculate_sentiments()



    def select_feeds(self):
        feed_manager = RSSFeedManager(self.master, self)
        self.master.wait_window(feed_manager)  # Wait for the window to be closed

    def manage_rss_feeds(self):
        RSSFeedManager(self.master, self)

    def manage_keywords(self):
        KeywordManager(self.master, self)
        self.recalculate_sentiments()

    def recalculate_sentiments(self):
        for item in self.news_data:
            item['sentiment'] = self.get_sentiment(item['title'] + " " + item['description'])
        self.update_treeview(self.news_data)

    def verify_rss_feed(self, url):
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            feed = feedparser.parse(response.content)
            if feed.entries:
                return True, f"Valid RSS feed. Found {len(feed.entries)} entries."
            else:
                return False, "RSS feed is valid but contains no entries."
        except requests.RequestException as e:
            return False, f"Error fetching RSS feed: {str(e)}"
        except Exception as e:
            return False, f"Error parsing RSS feed: {str(e)}"



    

class KeywordManager(ctk.CTkToplevel):
    def __init__(self, parent, news_app):
        super().__init__(parent)
        self.news_app = news_app
        self.title("Manage Keywords")
        self.geometry("600x400")

        # Make the window modal
        self.grab_set()
        self.transient(parent)
        self.focus_set()

        # Ensure the window stays on top
        self.attributes('-topmost', True)

        # Center the window on the screen
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        # Set color scheme
        self.configure(fg_color="white")

        self.create_widgets()

    def create_widgets(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self.positive_frame = ctk.CTkFrame(self.notebook, fg_color="white")
        self.negative_frame = ctk.CTkFrame(self.notebook, fg_color="white")

        self.notebook.add(self.positive_frame, text="Positive Keywords")
        self.notebook.add(self.negative_frame, text="Negative Keywords")

        self.create_keyword_list(self.positive_frame, self.news_app.positive_keywords, "positive")
        self.create_keyword_list(self.negative_frame, self.news_app.negative_keywords, "negative")

    def create_keyword_list(self, parent, keywords, keyword_type):
        listbox_frame = ctk.CTkFrame(parent, fg_color="white")
        listbox_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        listbox = tk.Listbox(listbox_frame, width=50, bg='white', fg='black')
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ctk.CTkScrollbar(listbox_frame, command=listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox.configure(yscrollcommand=scrollbar.set)

        for keyword in keywords:
            listbox.insert(tk.END, keyword)

        button_frame = ctk.CTkFrame(parent, fg_color="white")
        button_frame.pack(pady=10)

        ctk.CTkButton(button_frame, text="Add", command=lambda: self.add_keyword(keyword_type, listbox)).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Edit", command=lambda: self.edit_keyword(keyword_type, listbox)).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Delete", command=lambda: self.delete_keyword(keyword_type, listbox)).pack(side=tk.LEFT, padx=5)

    def add_keyword(self, keyword_type, listbox):
        dialog = ctk.CTkInputDialog(text=f"Enter a new {keyword_type} keyword:", title="Add Keyword")
        new_keyword = dialog.get_input()
        if new_keyword:
            logger.debug(f"Adding {keyword_type} keyword: {new_keyword}")
            if keyword_type == "positive":
                self.news_app.positive_keywords.append(new_keyword)
            else:
                self.news_app.negative_keywords.append(new_keyword)
            listbox.insert(tk.END, new_keyword)
            self.news_app.save_keywords()
            self.news_app.refresh_keywords()
            logger.debug(f"Updated {keyword_type} keywords: {self.news_app.positive_keywords if keyword_type == 'positive' else self.news_app.negative_keywords}")

    def edit_keyword(self, keyword_type, listbox):
        selected = listbox.curselection()
        if selected:
            index = selected[0]
            old_keyword = listbox.get(index)
            dialog = ctk.CTkInputDialog(text=f"Edit the {keyword_type} keyword:", title="Edit Keyword")
            new_keyword = dialog.get_input()
            if new_keyword:
                logger.debug(f"Editing {keyword_type} keyword: {old_keyword} -> {new_keyword}")
                if keyword_type == "positive":
                    self.news_app.positive_keywords[self.news_app.positive_keywords.index(old_keyword)] = new_keyword
                else:
                    self.news_app.negative_keywords[self.news_app.negative_keywords.index(old_keyword)] = new_keyword
                listbox.delete(index)
                listbox.insert(index, new_keyword)
                self.news_app.save_keywords()
                self.news_app.refresh_keywords()
                logger.debug(f"Updated {keyword_type} keywords: {self.news_app.positive_keywords if keyword_type == 'positive' else self.news_app.negative_keywords}")
        else:
            CTkMessagebox(title="No Selection", message="Please select a keyword to edit.", icon="warning")

    def delete_keyword(self, keyword_type, listbox):
        selected = listbox.curselection()
        if selected:
            index = selected[0]
            keyword = listbox.get(index)
            confirm = CTkMessagebox(
                title="Confirm Deletion",
                message=f"Are you sure you want to delete the keyword: {keyword}?",
                icon="warning",
                option_1="Yes",
                option_2="No",
                options=["Yes", "No"]
            )
            if confirm.get() == "Yes":
                logger.debug(f"Deleting {keyword_type} keyword: {keyword}")
                if keyword_type == "positive":
                    self.news_app.positive_keywords.remove(keyword)
                else:
                    self.news_app.negative_keywords.remove(keyword)
                listbox.delete(index)
                self.news_app.save_keywords()
                self.news_app.refresh_keywords()
                logger.debug(f"Updated {keyword_type} keywords: {self.news_app.positive_keywords if keyword_type == 'positive' else self.news_app.negative_keywords}")
                CTkMessagebox(title="Deletion Successful", message=f"Keyword '{keyword}' has been deleted.", icon="info")
        else:
            CTkMessagebox(title="No Selection", message="Please select a keyword to delete.", icon="warning")
            
class ConfirmDialog(ctk.CTkToplevel):
    def __init__(self, parent, title, message):
        super().__init__(parent)
        self.title(title)
        self.geometry("300x150")
        self.result = False

        # Make the window modal and keep it on top
        self.grab_set()
        self.transient(parent)
        self.focus_set()
        self.attributes('-topmost', True)

        # Center the window
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        ctk.CTkLabel(self, text=message).pack(pady=20)
        
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=10)
        
        ctk.CTkButton(button_frame, text="Yes", command=self.yes_click).pack(side=ctk.LEFT, padx=10)
        ctk.CTkButton(button_frame, text="No", command=self.no_click).pack(side=ctk.LEFT, padx=10)

    def yes_click(self):
        self.result = True
        self.destroy()

    def no_click(self):
        self.result = False
        self.destroy()



class RSSFeedManager(ctk.CTkToplevel):
    def __init__(self, parent, news_app):
        super().__init__(parent)
        self.news_app = news_app
        self.title("Manage RSS Feeds")
        self.geometry("1100x400")

        # Make the window modal and keep it on top
        self.grab_set()
        self.transient(parent)
        self.focus_set()
        self.attributes('-topmost', True)

        # Center the window
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        self.create_widgets()

    def create_widgets(self):
        listbox_frame = ctk.CTkFrame(self)
        listbox_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.listbox = ttk.Treeview(listbox_frame, columns=('Name', 'URL', 'Website'), show='headings')
        self.listbox.heading('Name', text='Feed Name')
        self.listbox.heading('URL', text='Feed URL')
        self.listbox.heading('Website', text='Website URL')
        self.listbox.column('Name', width=150)
        self.listbox.column('URL', width=300)
        self.listbox.column('Website', width=300)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = ctk.CTkScrollbar(listbox_frame, command=self.listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.listbox.configure(yscrollcommand=scrollbar.set)

        for feed in self.news_app.rss_feeds:
            self.listbox.insert('', 'end', values=(feed['name'], feed['url'], feed.get('website', '')))

        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=10)

        ctk.CTkButton(button_frame, text="Add", command=self.add_feed).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Edit", command=self.edit_feed).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Delete", command=self.delete_feed).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Verify", command=self.verify_feed).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Select All", command=self.select_all).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Deselect All", command=self.deselect_all).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Confirm Selection", command=self.confirm_selection).pack(side=tk.LEFT, padx=5)

    def add_feed(self):
        AddEditFeedDialog(self, self.news_app, mode='add')

    def edit_feed(self):
        selected = self.listbox.selection()
        if selected:
            item = self.listbox.item(selected[0])
            feed_name, feed_url, website_url = item['values']
            AddEditFeedDialog(self, self.news_app, mode='edit', initial_name=feed_name, initial_url=feed_url, initial_website=website_url)
        else:
            CTkMessagebox(title="No Selection", message="Please select a feed to edit.", icon="warning")


    def delete_feed(self):
        selected = self.listbox.selection()
        if selected:
            item = self.listbox.item(selected[0])
            values = item['values']
            
            # Ensure we have at least two values (name and URL)
            if len(values) < 2:
                CTkMessagebox(title="Error", message="Invalid feed data", icon="cancel")
                return
            
            feed_name, feed_url = values[:2]
            website = values[2] if len(values) > 2 else ''

            confirm_dialog = ConfirmDialog(self, "Confirm Deletion", f"Are you sure you want to delete the feed:\n{feed_name}?")
            self.wait_window(confirm_dialog)
            if confirm_dialog.result:
                self.listbox.delete(selected)
                # Update the rss_feeds list to remove the deleted feed
                self.news_app.rss_feeds = [feed for feed in self.news_app.rss_feeds 
                                           if feed['url'] != feed_url]
                self.news_app.save_rss_feeds()
                CTkMessagebox(title="Deletion Successful", message=f"Feed '{feed_name}' has been deleted.", icon="info")
        else:
            CTkMessagebox(title="No Selection", message="Please select a feed to delete.", icon="warning")


    def verify_feed(self):
        selected = self.listbox.selection()
        if selected:
            item = self.listbox.item(selected[0])
            values = item['values']
            
            # Ensure we have at least two values (name and URL)
            if len(values) < 2:
                CTkMessagebox(title="Error", message="Invalid feed data", icon="cancel")
                return
            
            feed_name, feed_url = values[:2]

            is_valid, message = self.news_app.verify_rss_feed(feed_url)
            if is_valid:
                CTkMessagebox(title="Valid RSS Feed", message=f"Feed '{feed_name}' is valid.\n\n{message}", icon="check")
            else:
                CTkMessagebox(title="Invalid RSS Feed", message=f"Feed '{feed_name}' is invalid.\n\n{message}", icon="cancel")
        else:
            CTkMessagebox(title="No Selection", message="Please select a feed to verify.", icon="warning")

    def select_all(self):
        for item in self.listbox.get_children():
            self.listbox.selection_add(item)

    def deselect_all(self):
        for item in self.listbox.get_children():
            self.listbox.selection_remove(item)

    def confirm_selection(self):
        selected = self.listbox.selection()
        self.news_app.selected_feeds = [self.listbox.item(item)['values'] for item in selected]
        CTkMessagebox(title="Selection Confirmed", message=f"Selected {len(self.news_app.selected_feeds)} feeds.", icon="info")
        self.after(100, self.destroy)  # Schedule destruction after a short delay

    def refresh_listbox(self):
        self.listbox.delete(*self.listbox.get_children())
        for feed in self.news_app.rss_feeds:
            self.listbox.insert('', 'end', values=(feed['name'], feed['url'], feed.get('website', '')))

class AddEditFeedDialog(ctk.CTkToplevel):
    def __init__(self, parent, news_app, mode='add', initial_name='', initial_url='', initial_website=''):
        super().__init__(parent)
        self.parent = parent
        self.news_app = news_app
        self.mode = mode
        self.title("Add Feed" if mode == 'add' else "Edit Feed")
        self.geometry("500x300")  # Increased size

        # Make the window modal
        self.grab_set()
        self.transient(parent)
        self.focus_set()

        # Ensure the window stays on top
        self.attributes('-topmost', True)

        # Center the window on the screen
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        self.create_widgets(initial_name, initial_url, initial_website)

    def create_widgets(self, initial_name, initial_url, initial_website):
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(padx=20, pady=20, fill=ctk.BOTH, expand=True)

        ctk.CTkLabel(main_frame, text="Feed Name:").grid(row=0, column=0, padx=5, pady=10, sticky='e')
        self.name_entry = ctk.CTkEntry(main_frame, width=300)  # Increased width
        self.name_entry.grid(row=0, column=1, padx=5, pady=10, sticky='ew')
        self.name_entry.insert(0, initial_name)

        ctk.CTkLabel(main_frame, text="Feed URL:").grid(row=1, column=0, padx=5, pady=10, sticky='e')
        self.url_entry = ctk.CTkEntry(main_frame, width=300)  # Increased width
        self.url_entry.grid(row=1, column=1, padx=5, pady=10, sticky='ew')
        self.url_entry.insert(0, initial_url)

        ctk.CTkLabel(main_frame, text="Website URL (optional):").grid(row=2, column=0, padx=5, pady=10, sticky='e')
        self.website_entry = ctk.CTkEntry(main_frame, width=300)  # Increased width
        self.website_entry.grid(row=2, column=1, padx=5, pady=10, sticky='ew')
        self.website_entry.insert(0, initial_website)

        button_frame = ctk.CTkFrame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, pady=20)

        ctk.CTkButton(button_frame, text="Save", command=self.save_feed).pack(side=ctk.LEFT, padx=10)
        ctk.CTkButton(button_frame, text="Cancel", command=self.destroy).pack(side=ctk.LEFT, padx=10)

        # Make the column with the entry widgets expandable
        main_frame.columnconfigure(1, weight=1)

    def save_feed(self):
        name = self.name_entry.get().strip()
        url = self.url_entry.get().strip()
        website = self.website_entry.get().strip()

        if not name or not url:
            CTkMessagebox(title="Invalid Input", message="Both name and URL are required.", icon="warning")
            return

        is_valid, message = self.news_app.verify_rss_feed(url)
        if not is_valid:
            CTkMessagebox(title="Invalid RSS Feed", message=message, icon="cancel")
            return

        if self.mode == 'add':
            self.news_app.rss_feeds.append({"name": name, "url": url, "website": website})
        else:
            for feed in self.news_app.rss_feeds:
                if feed['url'] == self.url_entry.get():
                    feed['name'] = name
                    feed['url'] = url
                    feed['website'] = website
                    break

        self.news_app.save_rss_feeds()
        self.parent.refresh_listbox()
        self.destroy()

class SelectFeedsDialog(ctk.CTkToplevel):
    def __init__(self, parent, news_app):
        super().__init__(parent)
        self.news_app = news_app
        self.title("Select Feeds")
        self.geometry("400x300")

        # Make the window modal
        self.grab_set()
        self.transient(parent)
        self.focus_set()

        # Ensure the window stays on top
        self.attributes('-topmost', True)

        # Center the window on the screen
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry('{}x{}+{}+{}'.format(width, height, x, y))

        self.create_widgets()

    def create_widgets(self):
        self.listbox = ttk.Treeview(self, columns=('Name', 'URL'), show='headings')
        self.listbox.heading('Name', text='Feed Name')
        self.listbox.heading('URL', text='Feed URL')
        self.listbox.column('Name', width=150)
        self.listbox.column('URL', width=200)
        self.listbox.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        for feed in self.news_app.rss_feeds:
            self.listbox.insert('', 'end', values=(feed['name'], feed['url']))

        button_frame = ctk.CTkFrame(self)
        button_frame.pack(pady=10)

        ctk.CTkButton(button_frame, text="Select All", command=self.select_all).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Deselect All", command=self.deselect_all).pack(side=tk.LEFT, padx=5)
        ctk.CTkButton(button_frame, text="Confirm Selection", command=self.confirm_selection).pack(side=tk.LEFT, padx=5)

    def select_all(self):
        for item in self.listbox.get_children():
            self.listbox.selection_add(item)

    def deselect_all(self):
        for item in self.listbox.get_children():
            self.listbox.selection_remove(item)

    def confirm_selection(self):
        selected = self.listbox.selection()
        self.news_app.selected_feeds = [self.listbox.item(item)['values'] for item in selected]
        messagebox.showinfo("Selection Confirmed", f"Selected {len(self.news_app.selected_feeds)} feeds.")
        self.destroy()



if __name__ == '__main__':
    multiprocessing.freeze_support()  # This is necessary for Windows
    root = ctk.CTk()
    app = NewsApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()